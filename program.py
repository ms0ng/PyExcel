import openpyxl
import os
import sys
import xml.etree.ElementTree as ET

walkDir = "Excel"
outputDir = "OutputData"

# excel表格的行列索引从1开始而非0
# 属性所在行
propertyRow = 2
# 第一个属性值所在列
valueColumn = 1


def WalkPath():
    for root, dirs, files in os.walk(walkDir):
        if(root.find(".svn") != -1):
            continue
        if(root.find(".git") != -1):
            continue
        for fName in files:
            fPath = os.path.join(root, fName)
            outputPath = os.path.join(root.replace(walkDir, outputDir), fName.replace("xlsx", "xml"))
            fPath = os.path.abspath(fPath)
            outputPath = os.path.abspath(outputPath)
            if not os.path.exists(os.path.dirname(outputPath)):
                os.mkdir(os.path.dirname(outputPath))
                open(outputPath, 'w+').close()
            HandleExcel(fPath, outputPath)
            print(fPath+"\t\t\t->\t\t"+outputPath+"\t...Done")


def HandleExcel(path, outputPath):
    wb = openpyxl.load_workbook(path)
    if(len(wb.sheetnames) < 0):
        return
    sheet = wb.worksheets[0]
    xml_Root = ET.Element("RECORDS")
    for i in range(propertyRow+1, sheet.max_row+1):
        xml_Node = ET.SubElement(xml_Root, "RECORD")
        for j in range(valueColumn, sheet.max_column+1):
            property = sheet.cell(propertyRow, j).value
            value = sheet.cell(i, j).value
            if(value == None):
                value = ''
            xml_Node.attrib[str(property)] = str(value)
    wb.close()
    prettyXml(xml_Root, '', '\n')
    xmlFile = ET.ElementTree(xml_Root)
    xmlFile.write(outputPath, "UTF-8")
    xmlFile.write(outputPath, encoding="utf-8", xml_declaration=True)



def prettyXml(element, indent, newline, level=0):
    if element:  
        if element.text == None or element.text.isspace():  
            element.text = newline + indent * (level + 1)
        else:
            element.text = newline + indent * \
                (level + 1) + element.text.strip() + \
                newline + indent * (level + 1)
    temp = list(element) 
    for subelement in temp:
        if temp.index(subelement) < (len(temp) - 1):
            subelement.tail = newline + indent * (level + 1)
            subelement.tail = newline + indent * level
        prettyXml(subelement, indent, newline, level=level + 1)

def main():
    os.chdir(os.path.dirname(sys.argv[0]))
    WalkPath()


if(__name__ == "__main__"):
    main()
